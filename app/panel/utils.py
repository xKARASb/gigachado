import datetime as dt
from typing import List, Set
from ..db.schemas.vacation import VacationRequestOutput as Vacation
from ..db.repos.deps import DepsRepository
from ..db.repos.employee import EmployeeRepository
from .schemas import VacationConflictResponse
import openpyxl as xl




def date_range(start: dt.date, end: dt.date) -> Set[dt.date]:
    return set([start + dt.timedelta(days=i) for i in range((end - start).days + 1)])

def get_overlaps(vacations: List[Vacation]) -> List[VacationConflictResponse]:
    conflicts = []
    for i in range(len(vacations)-1):
        for j in range(i+1, len(vacations)):
            first = date_range(vacations[i].start_at, vacations[i].end_at)
            second = date_range(vacations[j].start_at, vacations[j].end_at)
            overlap = first & second
            if len(overlap):
                conflicts.append(
                    VacationConflictResponse(
                        vacation1=vacations[i],
                        vacation2=vacations[j],
                        overlap=overlap,
                    )
                )
    return conflicts


def export_data_xl(vacations: List[Vacation], export_name, session):
    template = "./data/templates/vacations.xlsx"
    output = f"./data/exports/{export_name}.xlsx"
    start_row = 28
    start_col = 1
    
    workbook = xl.load_workbook(template)
    worksheet = workbook.active
    
    for i, vacation in enumerate(vacations):
        employee = EmployeeRepository(session).get_by_id(vacation.employee_id)
        dep = DepsRepository(session).get_by_id(vacation.dep_id)
        row = start_row + i
        worksheet.cell(row=row, column=start_col, value=i+1)
        worksheet.cell(row=row, column=start_col+1, value=dep.title)
        worksheet.cell(row=row, column=start_col+2, value=employee.post)
        worksheet.cell(row=row, column=start_col+3, value=f"{employee.name.capitalize()} {employee.last_name.capitalize()} {employee.patronymic.capitalize()}")
        worksheet.cell(row=row, column=start_col+4, value=employee.tabel_number)
        worksheet.cell(row=row, column=start_col+5, value=employee.vacation_days)
        worksheet.cell(row=row, column=start_col+6, value=employee.additional_days)
        worksheet.cell(row=row, column=start_col+7, value=employee.additional_days + employee.vacation_days)
        worksheet.cell(row=row, column=start_col+8, value=vacation.start_at)
    
    worksheet[f"B{len(vacations)+4}"] = "Начальник отдела по управлению персоналом"
    worksheet[f"J{len(vacations)+4}"] = "Т.В Свечникова"
    
    workbook.save(output)
    return output